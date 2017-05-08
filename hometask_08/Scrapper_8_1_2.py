
'''

'''

__author__ = 'Oleksandr Shapran'

import re
import requests
import json
import time
import asyncio
import psycopg2
import aiohttp
from lxml import html
import html as h
import logging
import csv
import pymongo
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, Alignment

logging.basicConfig(level=logging.INFO)

class Scrapper:
    def __init__(self, url, limit=2):
        self.url = url
        self.limit = limit
        self.results = []
        self.links = []
 
    async def __prepare(self):
        HEADERS = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'accept-encoding': 'gzip, deflate, sdch, br',
            'accept-language': 'ru-RU,ru;q=0.8,en-US;q=0.6,en;q=0.4',
            'cookie': '__cfduid=d411cccf4e5f1772e67bfe5e7fcfcddb51493969528; __gads=ID=d549c4e37012c484:T=1493969538:S=ALNI_MZnB1bQ6C74zWdtY-3y6MBT0MD2rg; _ga=GA1.2.1992315233.1493969540; _gid=GA1.2.1201668070.1493972828',
            'if-modified-since': 'Fri, 05 May 2017 08:25:12 GMT',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.96 Mobile Safari/537.36',
           # ':authority': 'coinmarketcap.com',
           # ':method': 'GET',
           # ':path': '/all/views/all/',
           # ':scheme': 'https',
            }

        conn = aiohttp.TCPConnector(verify_ssl=True)
        self.session = aiohttp.ClientSession(connector=conn, headers=HEADERS)
        return

    def start(self):
        ''' create loop for scrapping'''
        event_loop = asyncio.get_event_loop()
        try:
            event_loop.run_until_complete(self.__prepare())
            event_loop.run_until_complete(self.__run())
            
        finally:
            self.session.close()
            event_loop.close()
        

    async def __run(self):
        ''' Crawl URL '''
        links = [self.url]

        semaphore = asyncio.BoundedSemaphore(self.limit)
        tasks = []
        results = []

        for link in links:
            tasks.append(self.crawl(link, semaphore))

        for task in asyncio.as_completed(tasks):
            result = await task
            results.append(result)
            task.close()

        return results


    async def crawl(self, url, semaphore):
        ''' Crawl URL '''
        async with semaphore:

            resp = await self.session.get(url)

            if resp.status == 200:
                page = await resp.text()
                root = html.fromstring(page)

                currencies = root.xpath('//table[@id="currencies-all"]/tbody/tr')
                for currency in currencies:
                    item = []
                    try:
                        #position
                        position = int(self.get_xpath_value(currency, './td', 0, './text()'))
                        item.append(position)
                        
                        #name
                        name = self.get_xpath_value(currency, './td', 1, './a/text()')
                        item.append(name)
                        #symbol
                        symbol = self.get_xpath_value(currency, './td', 2, './text()')
                        item.append(symbol)
                        #market_cap
                        market_cap = self.get_xpath_value(currency, './td', 3, './text()')
                        item.append(market_cap)
                        #price
                        price = self.get_xpath_value(currency, './td', 4, './a/text()')
                        item.append(price)
                        #supply
                        supply = self.get_xpath_value(currency, './td', 5, './descendant-or-self::text()', -1)
                        item.append(supply)
                        #volume
                        volume = self.get_xpath_value(currency, './td', 6, './a/text()')
                        item.append(volume)
                        #h1
                        h1 = self.get_xpath_value(currency, './td', 7, './text()')
                        item.append(h1)
                        #h24
                        h24 = self.get_xpath_value(currency, './td', 8, './text()')
                        item.append(h24)
                        #d7
                        d7 = self.get_xpath_value(currency, './td', 9, './text()')
                        item.append(d7)
                        #appends to results
                        self.results.append(item)

                       ## print("{} -> {}-> {}-> {}-> {}-> {}-> {}-> {}-> {}-> {}-> {}".format(position, name, symbol,
                         #   #                    market_cap, price, supply,
                        # #                       volume, h1, h24, d7)) 
                    except Exception as e:
                        print(e)
                return

    def get_xpath_value(self, element, selecto1, number1, selector2, number2=0):
        try:
            if number2 != -1:
                return element.xpath(selecto1)[number1].xpath(selector2)[number2].strip()
            else:
                list_inner_text = element.xpath(selecto1)[number1].xpath(selector2)
                return ''.join( map((lambda x: x.strip()), list_inner_text) )
        except Exception as e:
            ##print("{} -> {}-> {}-> {} ".format(selecto1, number1, selector2, number2))
            print(e)
            return ''

 
    def save_to_xlsx(self):
        ''' Save data to formatted xlsx file'''
        #output file name and path
        file = "output.xlsx"

        wb = Workbook()
        # grab the active worksheet
        wsheet = wb.active
        
        #set styles for columns 
        RED = Font(color=colors.RED)
        GREEN = Font(color=colors.GREEN)
        BLUE = Font(color=colors.BLUE)
        BLACK = Font(color=colors.BLACK)    
        RERSENT_FORMAT = '0.00%;[Red]-0.00%'
        DOLLAR_FORMAT = '[$$-409]#,##0;[RED]-[$$-409]#,##0'
        DOLLAR_FORMAT_EXT = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
        NUMBER_FORMAT = '# ##0'
        DEFAULT_FORMAT = "@"
        
        colunm_style = {
            "B": {'color': BLUE, 'format': DEFAULT_FORMAT, 'width': 25 },
            "D": {'color': BLACK, 'format': DOLLAR_FORMAT, 'width': 25 },
            "E": {'color': BLUE, 'format': DOLLAR_FORMAT_EXT },
            "F": {'color': BLUE, 'format': NUMBER_FORMAT, 'width': 25 },
            "G": {'color': BLUE, 'format': DOLLAR_FORMAT, 'width': 25 },
            "H": {'color': GREEN, 'format': RERSENT_FORMAT },
            "I": {'color': GREEN, 'format': RERSENT_FORMAT },
            "J": {'color': GREEN, 'format': RERSENT_FORMAT }
            }
        for k, v in colunm_style.items():
            col = wsheet.column_dimensions[k]
            col.font = v['color']
            col.number_format = v['format']
            if 'width' in v:
                col.width = v['width']

        #set style for header
        wsheet.row_dimensions[1].font = BLACK
        wsheet.row_dimensions[1].alignment=Alignment(horizontal='center')
        
        #create header
        column_header = ['#', 'Name', 'Symbol', 'Market Cap', 'Price', \
                         'Circulating Supply', 'Volume (24h)', '% 1h', '% 24h', '% 7d']
        for x in range(1, len(column_header)+1):
            wsheet.cell(row = 1, column = x).value = column_header[x-1]
            
     
        #fill table with scrapped data
        for y in range(0, len(self.results)):
            for x in range(0, len(self.results[0])):
                value = self.results[y][x]
                if x >= 3:
                    value = self.get_number(value)
                wsheet.cell(row = y+2, column = x+1).value = value

        # Save the file
        wb.save(file)
        print('File saved!!!')

    def get_number(self, string):
        try:            
            string_cleared = string.replace('$', '').replace('%', '').replace(',', '')
            string_cleared = float(string_cleared)
            return string_cleared/100 if string.find('%')>0 else string_cleared
        except:
            return 0.0

    
if __name__ == "__main__":
    start_time = time.time()
    ##
    print("Start")
    URL = 'https://coinmarketcap.com/all/views/all/'
    scrapper = Scrapper(URL, limit=2)
    scrapper.start()
    scrapper.save_to_xlsx()
    print("--- %s seconds ---" % (time.time() - start_time))
 
     

    
    
  
